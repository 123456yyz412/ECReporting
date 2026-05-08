from __future__ import annotations

import csv
import io
import re
from typing import Any, Iterable

import openpyxl
from openpyxl.workbook import Workbook
from psycopg import sql as psql

from apps.datasources.models import DataSourceType
from apps.datasources.services import connect, get_table_columns

from .models import UploadModule


_ident_re = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


def _mysql_ident(name: str) -> str:
    if not _ident_re.match(name):
        raise ValueError("invalid identifier")
    return f"`{name}`"


def build_template_xlsx(module: UploadModule) -> bytes:
    cols = get_table_columns(module.datasource, table_name=module.table_name, schema=module.schema or None)
    wb: Workbook = openpyxl.Workbook()
    ws = wb.active
    ws.title = module.name[:31]
    ws.append([c.name for c in cols])
    ws.append([module.columns.get(c.name) or (c.comment or "") for c in cols])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_template_csv(module: UploadModule) -> bytes:
    cols = get_table_columns(module.datasource, table_name=module.table_name, schema=module.schema or None)
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow([c.name for c in cols])
    writer.writerow([module.columns.get(c.name) or (c.comment or "") for c in cols])
    return out.getvalue().encode("utf-8-sig")


def _iter_rows_from_csv(content: bytes) -> tuple[list[str], list[list[Any]]]:
    text = content.decode("utf-8-sig", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if len(rows) < 2:
        raise ValueError("CSV需要至少两行表头（英文名/中文名）")
    header = [h.strip() for h in rows[0]]
    data_rows = rows[2:]
    return header, data_rows


def _iter_rows_from_xlsx(content: bytes) -> tuple[list[str], list[list[Any]]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("XLSX需要至少两行表头（英文名/中文名）")
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    data_rows = [list(r) for r in rows[2:]]
    return header, data_rows


def parse_upload_file(filename: str, content: bytes) -> tuple[list[str], list[list[Any]]]:
    fn = filename.lower()
    if fn.endswith(".csv"):
        return _iter_rows_from_csv(content)
    if fn.endswith(".xlsx"):
        return _iter_rows_from_xlsx(content)
    raise ValueError("仅支持CSV或XLSX文件")


def insert_rows(module: UploadModule, header: list[str], rows: list[list[Any]], *, max_rows: int = 20000) -> int:
    header = [h.strip() for h in header if h and h.strip()]
    if not header:
        raise ValueError("缺少列名")
    if len(rows) > max_rows:
        raise ValueError(f"单次最多上传{max_rows}行")

    clean_rows: list[list[Any]] = []
    for r in rows:
        if r is None:
            continue
        r = list(r)
        if all((c is None or str(c).strip() == "") for c in r):
            continue
        r = r[: len(header)]
        r += [None] * (len(header) - len(r))
        clean_rows.append([None if (isinstance(c, str) and c.strip() == "") else c for c in r])

    if not clean_rows:
        return 0

    if module.datasource.db_type == DataSourceType.POSTGRES:
        schema = (module.schema or "").strip() or (module.datasource.schema or "").strip() or "public"
        with connect(module.datasource) as conn:
            with conn.cursor() as cur:
                cur.execute("begin")
                table_ident = psql.Identifier(schema, module.table_name)
                cols_ident = [psql.Identifier(c) for c in header]
                insert_sql = psql.SQL("insert into {} ({}) values ({})").format(
                    table_ident,
                    psql.SQL(",").join(cols_ident),
                    psql.SQL(",").join([psql.Placeholder()] * len(header)),
                )
                inserted = 0
                for i in range(0, len(clean_rows), 1000):
                    chunk = clean_rows[i : i + 1000]
                    cur.executemany(insert_sql, chunk)
                    inserted += len(chunk)
                cur.execute("commit")
        return inserted

    if module.datasource.db_type == DataSourceType.MYSQL:
        schema = module.datasource.database
        table = _mysql_ident(module.table_name)
        cols = ",".join(_mysql_ident(c) for c in header)
        placeholders = ",".join(["%s"] * len(header))
        insert_sql = f"insert into {table} ({cols}) values ({placeholders})"
        with connect(module.datasource) as conn:
            with conn.cursor() as cur:
                inserted = 0
                for i in range(0, len(clean_rows), 1000):
                    chunk = clean_rows[i : i + 1000]
                    cur.executemany(insert_sql, chunk)
                    inserted += len(chunk)
        return inserted

    raise ValueError("Unsupported db_type")

